# -*- coding: utf-8 -*-
#Written in Python 3.6
#Python 2 compliant as of 6/16/17 
#Written in openpyxl 2.4, should work for any previous 2.x release

"""
Python functions for the Factorial Mass ERP Univariate Toolbox

AUTHOR: Eric Fields
VERSION DATE: 26 April 2018
"""

import sys
import re
import openpyxl
if int(openpyxl.__version__[0]) < 2:
    raise RuntimeError('openpyxl must be version 2.0 or greater.')
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule

if sys.version_info.major == 2:
    input = raw_input
    range = xrange

def format_xls(spreadsheet):
    """
    Take spreadsheet output from FMUT functions and apply formatting
    """
    
    wb = openpyxl.load_workbook(spreadsheet)
    
    #Delete blank sheets
    for sheet_name in wb.get_sheet_names():
        if re.match('Sheet\d', sheet_name):
            sheet2remove = wb.get_sheet_by_name(sheet_name)
            wb.remove_sheet(sheet2remove)
    wb.active = 0
    
    #Define some fill styles for use below
    whiteFill = PatternFill(start_color='ffffff', end_color='ffffff', fill_type='solid')
    yellowHighlight = PatternFill(start_color='f7ff23', end_color='f7ff23', fill_type='solid')
    
    for sheet_name in wb.get_sheet_names():
        
        sheet = wb.get_sheet_by_name(sheet_name)
        
        #Test summary sheet
        if sheet_name == 'test summary':
            sheet.column_dimensions['A'].width = 40
            sheet.column_dimensions['B'].width = 200
            for cell in sheet['B']:
                cell.alignment = Alignment(horizontal='left')
            #Format critical values
            for row in range(1, sheet.max_row+1):
                if 'critical value' in sheet['A'+str(row)].value:
                    sheet['B'+str(row)].number_format = '0.00'
        
        #cluster_summary sheet
        if sheet_name == 'cluster summary':
            #Set column dimensions
            for i in range(1, sheet.max_column+1):
                if i % 3 == 1:
                    sheet.column_dimensions[get_column_letter(i)].width = 19
                elif i % 3 == 2:
                    sheet.column_dimensions[get_column_letter(i)].width = 10
                elif i % 3 == 0:
                    sheet.column_dimensions[get_column_letter(i)].width = 5
            #Freeze header for easier viewing
            sheet.freeze_panes = sheet['A2']
            #Set header style
            for cell in sheet[1]:
                cell.font = Font(sz=13, bold=True)
            #Format decimal places of stats
            for row in range(2, sheet.max_row+1):
                if any(cell.value=='cluster mass' for cell in sheet[str(row)]):
                    for cell in sheet[row]:
                        cell.number_format = '0.00'
                elif any(cell.value=='p-value' for cell in sheet[str(row)]):
                    for cell in sheet[row]:
                        cell.number_format = '0.000'
            #Left align everything for easier viewing
            for col in range(2, sheet.max_column+1):
                if col % 3 == 2:
                    for row in range(3, sheet.max_row+1):
                        sheet.cell(row=row, column=col).alignment = Alignment(horizontal='left')
            #Bold and highlight significant clusters
            for row in range(1, sheet.max_row+1):
                for cell in sheet[row]:
                    if cell.value=='p-value' and cell.offset(row=0, column=1).value <= .05:
                        cell.offset(row=0, column=1).fill = yellowHighlight
                        for r in range(cell.row-2, cell.row+7):
                            for c in range(column_index_from_string(cell.column), column_index_from_string(cell.column)+2):
                                sheet.cell(row=r, column=c).font = Font(bold=True)
                        
                        
        #Some general stuff for remaining sheets
        if sheet_name.endswith('clust_IDs') or sheet_name.endswith('pvals') or sheet_name.endswith('_obs'):
            #Make headers bold
            for cell in sheet[1]:
                cell.font = Font(bold=True)
                cell.number_format = '0'
            for cell in sheet['A']:
                cell.font = Font(bold=True)
            #Freeze headers for easier viewing
            sheet.freeze_panes = sheet['B2']
        
        #Format cluster ID sheets
        if sheet_name.endswith('clust_IDs'):
            #Reduce column width
            for i in range(1, sheet.max_column+1):
                sheet.column_dimensions[get_column_letter(i)].width = 4
            #Apply a different color to each cluster
            max_cell = get_column_letter(sheet.max_column) + str(sheet.max_row)
            clust_colors = ('36ec41', '003fbb', 'b4b500', 'c777ff', '00f7b8', '2f0067', '537e00', 'ff6dbf', '8befff', 'ff5227', '0171aa', 'a92900', '00727b', 'b60070', 'faffd1', '540042', 'a96500', 'e2caff', 'ffa177','6a001f')
            for clust in range(20):
			    color_idx = clust % 20
                clustFill = PatternFill(start_color=clust_colors[color_idx], end_color=clust_colors[color_idx], fill_type='solid')
                sheet.conditional_formatting.add('B2:'+max_cell,
                        CellIsRule(operator='equal', formula=[clust+1], stopIfTrue=True, fill=clustFill))
            #Reduce font size and clear locations not included in cluster
            for row in range(2, sheet.max_row+1):
                for cell in sheet[row]:
                    if column_index_from_string(cell.column) > 1:
                        cell.font = Font(sz=10)
                    if not cell.value:
                        cell.value = None
            
        #Format t-obs F-obs sheets
        elif sheet_name.endswith('_obs'):
            #Reduce column width
            for i in range(1, sheet.max_column+1):
                sheet.column_dimensions[get_column_letter(i)].width = 6
            #Format numbers and apply conditional formatting
            max_cell = get_column_letter(sheet.max_column) + str(sheet.max_row)
            data_range = 'B2:'+max_cell
            for row in range(2, sheet.max_row+1):
                for cell in sheet[row]:
                    if cell.column != 'A':
                        cell.number_format = '0.00'
                        cell.font = Font(sz=10)
            if sheet_name.endswith('F_obs'):
                sheet.conditional_formatting.add(data_range,
                                ColorScaleRule(start_type='num', start_value=1, start_color='ffffff',
                                               end_type='max',   end_color='448452'))
            elif sheet_name.endswith('t_obs'):
                sheet.conditional_formatting.add(data_range,
                                ColorScaleRule(start_type='min', start_color='0047ba',
                                               mid_type='num', mid_value=0, mid_color='ffffff',
                                               end_type='max',   end_color='c63535'))
        
        #Format p-value sheets
        elif sheet_name.endswith('_pvals'):
            #Reduce column width
            for i in range(2, sheet.max_column+1):
                sheet.column_dimensions[get_column_letter(i)].width = 6
            #Format numbers and apply conditional formatting
            max_cell = get_column_letter(sheet.max_column) + str(sheet.max_row)
            data_range = 'B2:'+max_cell
            for row in range(2, sheet.max_row+1):
                for cell in sheet[row]:
                    if cell.column != 'A':
                        cell.number_format = '0.000'
                        cell.font = Font(sz=10)
            sheet.conditional_formatting.add(data_range,
                            CellIsRule(operator='greaterThan', formula=[0.05], stopIfTrue=True, fill=whiteFill))
            sheet.conditional_formatting.add(data_range,
                            ColorScaleRule(start_type='num', start_value=0,  start_color='4e875b',
                                           end_type='num',   end_value=0.05, end_color='c9ffd5'))
    
    #Save
    wb.save(spreadsheet)

if __name__ == '__main__':
    if len(sys.argv) > 3:
        raise RuntimeError('fmut.py can only take one argument-option pair.')
    elif '--format_xls' in sys.argv:
        format_xls(sys.argv[sys.argv.index('--format_xls')+1])
    else:
        spreadsheet = input('To format an FMUT output spreadsheet, give the filename with the full path: ')
        format_xls(spreadsheet)
